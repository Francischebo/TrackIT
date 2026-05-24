from datetime import datetime
import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.utils.formatted_export import TRACKIT_BRAND, _format_generated_at


class ProfessionalPDFReport:
    """Generate institutional PDF reports: metadata header, then column headings and rows."""

    def __init__(self, title, org_name, landscape_mode=False, subtitle=None):
        self.buffer = io.BytesIO()
        self.pagesize = landscape(A4) if landscape_mode else A4
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=self.pagesize,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
        )
        self.title = title
        self.org_name = org_name
        self.subtitle = subtitle
        self.styles = getSampleStyleSheet()
        self.elements = []
        self._table_payload = None

        self.title_style = ParagraphStyle(
            "ReportTitle",
            parent=self.styles["Heading1"],
            fontSize=22,
            spaceAfter=6,
            textColor=colors.HexColor("#1e293b"),
        )
        self.meta_style = ParagraphStyle(
            "MetaStyle",
            fontSize=9,
            textColor=colors.HexColor("#64748b"),
            spaceAfter=4,
        )

    def _build_header_elements(self):
        block = [
            Paragraph(TRACKIT_BRAND, self.meta_style),
            Paragraph(self.title, self.title_style),
            Paragraph(f"<b>Organization:</b> {self.org_name}", self.meta_style),
        ]
        if self.subtitle:
            block.append(Paragraph(self.subtitle, self.meta_style))
        block.append(
            Paragraph(f"<b>Generated:</b> {_format_generated_at()}", self.meta_style)
        )
        block.append(Spacer(1, 10 * mm))
        return block

    def create_table(self, data, col_widths=None):
        self._table_payload = (data, col_widths)

    def _append_table(self):
        if not self._table_payload:
            return
        data, col_widths = self._table_payload
        if not data:
            return

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("TOPPADDING", (0, 0), (-1, 0), 10),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#334155")),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f8fafc")],
                    ),
                ]
            )
        )
        self.elements.append(table)

    def generate(self):
        self.elements = self._build_header_elements() + self.elements
        self._append_table()
        self.doc.build(self.elements)
        self.buffer.seek(0)
        return self.buffer


class ProfessionalExcelReport:
    """Excel exports with title block (rows 1–5), headings row 6, data from row 7."""

    HEADER_ROW = 6
    DATA_START_ROW = 7

    @staticmethod
    def _populate_sheet(ws, title, headers, data, org_name=None, subtitle=None):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        title_font = Font(bold=True, size=16, color="1E293B")
        meta_font = Font(size=10, color="64748B")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        last_col = max(len(headers), 1)
        last_letter = get_column_letter(last_col)

        ws["A1"] = TRACKIT_BRAND
        ws["A1"].font = meta_font
        ws.merge_cells(f"A1:{last_letter}1")
        ws["A2"] = title
        ws["A2"].font = title_font
        ws.merge_cells(f"A2:{last_letter}2")
        meta_row = 3
        if org_name:
            ws[f"A{meta_row}"] = f"Organization: {org_name}"
            ws[f"A{meta_row}"].font = meta_font
            ws.merge_cells(f"A{meta_row}:{last_letter}{meta_row}")
            meta_row += 1
        if subtitle:
            ws[f"A{meta_row}"] = subtitle
            ws[f"A{meta_row}"].font = meta_font
            ws.merge_cells(f"A{meta_row}:{last_letter}{meta_row}")
            meta_row += 1
        ws[f"A{meta_row}"] = f"Generated: {_format_generated_at()}"
        ws[f"A{meta_row}"].font = meta_font
        ws.merge_cells(f"A{meta_row}:{last_letter}{meta_row}")

        header_row = ProfessionalExcelReport.HEADER_ROW
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        data_start = ProfessionalExcelReport.DATA_START_ROW
        for row_offset, row_data in enumerate(data):
            row_num = data_start + row_offset
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        for col_num, header in enumerate(headers, 1):
            letter = get_column_letter(col_num)
            max_len = len(str(header))
            for row_data in data[:300]:
                if col_num <= len(row_data):
                    max_len = max(max_len, len(str(row_data[col_num - 1] or "")))
            ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 45)

        ws.freeze_panes = f"A{data_start}"

    @staticmethod
    def generate(title, headers, data, org_name=None, subtitle=None, sheet_title=None):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet_title or title or "Report")[:31]
        ProfessionalExcelReport._populate_sheet(
            ws, title, headers, data, org_name=org_name, subtitle=subtitle
        )
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_multi_sheet(org_name, report_title, sheets):
        """sheets: [{title, headers, rows, sheet_title?, subtitle?}, ...]"""
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for spec in sheets:
            name = (spec.get("sheet_title") or spec["title"])[:31]
            ws = wb.create_sheet(title=name)
            ProfessionalExcelReport._populate_sheet(
                ws,
                spec["title"],
                spec["headers"],
                spec["rows"],
                org_name=org_name,
                subtitle=spec.get("subtitle"),
            )
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
